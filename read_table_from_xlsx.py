import openpyxl

def read_two_column_table_from_xlsx(filepath):
    """
    Opens an XLSX file, reads a 2-column table from the first sheet,
    and returns it as a 2-dimensional list.

    Args:
        filepath (str): The path to the XLSX file.

    Returns:
        list: A 2-dimensional list representing the 2-column table.
              Returns an empty list if the file is not found or if
              there's an issue reading the data.
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active  # Get the first (active) sheet

        data = []
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            # values_only=True ensures we get the cell values directly
            # max_col=2 ensures we only read up to the second column
            if len(row) == 2:  # Ensure the row has exactly two columns of data
                data.append(list(row))
            elif len(row) == 1 and row[0] is not None:
                # Handle cases where the second column might be empty but the first isn't
                data.append([row[0], ""])
            # Rows with no data in the first two columns (e.g., empty rows) will be skipped

        return data[1:0]

    except FileNotFoundError:
        print(f"Error: The file '{filepath}' was not found.")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

# --- Example Usage ---
if __name__ == "__main__":
    # Create a dummy XLSX file for testing
    try:
        workbook_test = openpyxl.Workbook()
        sheet_test = workbook_test.active
        sheet_test.title = "Sheet1"

        # Add some data
        sheet_test['A1'] = 'Header 1'
        sheet_test['B1'] = 'Header 2'
        sheet_test['A2'] = 'Value A2'
        sheet_test['B2'] = 'Value B2'
        sheet_test['A3'] = 'Value A3'
        sheet_test['B3'] = 'Value B3'
        sheet_test['A4'] = 'Only Col 1' # Test case for missing second column data
        sheet_test['A5'] = None        # Test case for empty row (should be skipped)
        sheet_test['B5'] = None
        sheet_test['A6'] = 'Value A6'
        sheet_test['B6'] = 'Value B6'

        test_file = "my_test_table.xlsx"
        workbook_test.save(test_file)
        print(f"Created a dummy XLSX file: {test_file}")

        # Read the table from the dummy file
        table_data = read_two_column_table_from_xlsx(test_file)

        if table_data:
            print("\nTable data from the XLSX file:")
            for row in table_data:
                print(row)
        else:
            print("\nNo data was read or an error occurred.")

    except Exception as e:
        print(f"Error during example usage: {e}")